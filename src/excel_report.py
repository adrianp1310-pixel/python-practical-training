from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def save_report(df: pd.DataFrame, path: Path) -> bool:
    """
    Zapisuje DataFrame do pliku Excel.

    Args:
        df: DataFrame który chcemy zapisać.
        path: Ścieżka gdzie ma być zapisany plik w .xlsx.

    Returns:
        True jeśli zapis się udał. False jeśli wystąpił błąd.
    """
    try:
        df.to_excel(path, index=False)
        return True
    except OSError:
        return False


def style_header(path: Path) -> bool:
    """
    Stylizuje pierwszy wiersz czyli nagłówek w pliku Excel.
    Nagłówek otrzymuje:
    - pogrubioną białą czcionkę
    - niebieskie tło
    - wyśrodkowanie tekstu

    Args:
        path: Ścieżka do istniejącego pliku Excel.

    Returns:
        True jeśli stylizacja się udała. False jeśli błąd.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        font = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="4472C4")
        alignment = Alignment(horizontal="center")
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
        wb.save(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False


def set_column_widths(path: Path, widths: dict[str, int]) -> bool:
    """
    Ustawia szerokość kolumn w pliku Excel.

    Args:
        path: Ścieżka do istniejącego pliku Excel.
        widths: Słownik z szerokościami kolumn.

    Returns:
        True jeśli ustawienie się udało. False jeśli błąd.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        wb.save(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False

def freeze_header(path: Path) -> bool:
    """
    Zamraża pierwszy wiersz w pliku Excel.

    Args:
        path: Ścieżka do pliku Excel.

    Returns:
        Zwraca True jeśli zamrożenie się udało. Zwraca False jeśli wystąpi błąd
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        ws.freeze_panes = "A2"
        wb.save(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False


def stripe_rows(path: Path, color: str="D9E2F3") -> bool:
    """
    Zmienia kolor parzystych wierszy w pliku Excel.

    Args:
        path: Ścieżka do pliku Excel.
        color: Kolor parzystych wierszy w pliku. Domyślnie ustawiony na D9E2F3.

    Returns:
        Zwraca True jeśli kolor został zmieniony. Zwraca False jeśli
        wystąpi błąd
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        fill = PatternFill("solid", fgColor=color)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = fill
        wb.save(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False


def add_borders(path: Path, style: str="thin", color: str="000000") -> bool:
    """
    Zmienia obramowanie komórek w pliku Excel.

    Args:
        path: Ścieżka do pliku Excel.
        style: Definiuje grubość obramowania komórki.
        color: Definiuje kolor obramowania komórki.
    Returns:
        Zwraca True jeśli obramowanie zostało zmienione. Zwraca False jeśli
        wystąpi błąd
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        side = Side(style=style, color=color)
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.border = border
        wb.save(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False

def auto_fit_columns(path: Path, padding: int=2) -> bool:
    """
    Automatycznie dopasowuje szerokość komórek w pliku Excel.

    Args:
        path: Ścieżka do pliku Excel.
        padding: Dodatkowa przestrzeń po obu stronach komórki. Domyślnie
        ustawiona na 2.
    Returns:
        Zwraca True jeśli auto-dopasowanie zostało ustawione. Zwraca False
        jeśli wystąpi błąd.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            letter = column_cells[0].column_letter
            ws.column_dimensions[letter].width = max_length + padding
        wb.save(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False




if __name__ == "__main__":
    df = pd.DataFrame([
        {"product": "Laptop Dell", "category": "Elektronika",
         "revenue": 17500},
        {"product": "Monitor", "category": "Elektronika",
         "revenue": 12600},
    ])
    path = Path("test_report.xlsx")
    result = save_report(df, path)
    print(result)
    result = style_header(path)
    print(result)
    result = set_column_widths(path, {"A": 20, "B": 15, "C": 12})
    print(result)
    result = freeze_header(path)
    print(result)
    result = stripe_rows(path, "D9E2F3")
    print(result)
    result = add_borders(path, "thin", "000000")
    print(result)
    result = auto_fit_columns(path)
    print(result)
